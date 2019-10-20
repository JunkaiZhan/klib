#!/usr/bin/python3

# Author : zhanjunkai
# Data   : 2019/09/08

import re
import sys
import os
import openpyxl
import subprocess


# =============================================
# Function    : sh(cmd)
# Description : do cmd with bash shell
# ============================================= 
def sh(cmd):
    subprocess.call(cmd, shell=True)

# =============================================
# Function    : check_file(file_name)
# Description : check if the file exists
# ============================================= 
def check_file (file_name):
    if os.path.exists(file_name):
        print("The file name is : " + file_name)
    else:
        print("The file: " + file_name + " is not found !!")
        sys.exit(0)

# =====================================================
# Function    : check_empty(obj)
# Description : check if the list is empty
# ===================================================== 
def check_empty(obj):
    if (len(obj) == 0):
        print("The return object is empty, please check the pattern or file")

# ====================================================================
# Function    : excel_int2letter(int)
# Description : transform the int number to letter for column of excel
# ====================================================================
def excel_int2letter(int):
    return openpyxl.cell.get_column_letter(int)

# ====================================================================
# Function    : excel_letter2int(letter)
# Description : transform the letter to int number for column of excel
# ====================================================================
def excel_letter2int(letter):
    return openpyxl.cell.column_index_from_string(letter)

# =====================================================
# Function    : get_list_from_file(file_name, pattern)
# Description : get a queue of elements that are
#               matched in the file
# =====================================================
def get_list_from_file(file_name, pattern):
    check_file(file_name)
    file        = open(file_name, 'r')
    content     = file.read()
    return_list = re.findall(content, pattern, re.I|re.M)
    check_empty(return_list)
    file.close()
    return return_list

# =====================================================
# Function    : get_list_from_string(pool, pattern)
# Description : get a queue of elememts that are 
#               matched in the string pool
# ===================================================== 
def get_list_from_string(pool, pattern):
    return_list = re.findall(pool, pattern, re.I|re.M)
    check_empty(return_list)
    return return_list

# ============================================================================
# Function    : get_dict_from_file(file_name, key_pattern, value_pattern)
# Description : get a dictionary which map the key_pattern to 
#               the value_patten in the file foreach line
# ============================================================================
def get_dict_from_file(file_name, key_pattern, value_pattern):
    check_file(file_name)
    file        = open(file_name, "r")
    content     = file.readlines()
    dict        = {}
    for line in content:
        key       = re.search(key_pattern, line, re.I|re.M)
        value     = re.search(value_pattern, line, re.I|re.M)
        if (key and value):
            dict[key.group()] =  value.group()
    file.close()
    check_empty(dict)
    return dict

# =====================================================
# Function    : get_sheet_from_file(file_name, sheet_name)
# Description : get the sheet handle from excel file
# ===================================================== 
def get_sheet_from_file(file_name, sheet_name):
    check_file(file_name)
    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]
    return ws

# =====================================================
# Function    : get_cell_from_sheet(sheet, row, column)
# Description : get the cell value from the work sheet
# ===================================================== 
def get_cell_from_sheet(sheet, row, column):
    cell_value = sheet.cell(row=row, column=column).value
    return cell_value

# ===========================================================================
# Function    : get_row_from_sheet(sheet, row_index, column_start, column_end)
# Description : get the list of values in row_index from column_start
#               to column_end
# ===========================================================================
def get_row_from_sheet(sheet, row_index, column_start, column_end):
    row_list = []
    for _ in range(column_start, column_end+1):
        cell_value = get_cell_from_sheet(sheet, row_index, _)
        row_list.append(cell_value)
    check_empty(row_list)
    return row_list

# ===========================================================================
# Function    : get_column_from_sheet(sheet, column_index, row_start, row_end)
# Description : get the list of values in column_index from row_start 
#               to row_end
# ===========================================================================
def get_column_from_sheet(sheet, column_index, row_start, row_end):
    column_list = []
    for _ in range(row_start, row_end+1):
        cell_value = get_cell_from_sheet(sheet, _, column_index)
        column_list.append(cell_value)
    check_empty(column_list)
    return column_list

# ====================================================================
# Function    : get_index_in_row_header(sheet, item_name, header_pos = 1)
# Description : get the index of the item in the header
# ====================================================================
def get_index_in_row_header (sheet, item_name, header_pos = 1):
    index    = -1
    max_col  = sheet.max_column
    row_list = get_row_from_sheet(sheet, header_pos, 1, max_col)
    for _ in range(0, len(row_list)):
        if (item_name == row_list[_]):
            index = _ + 1
            return index
    print("The item: " + item_name + " is not found in the header, please check the item name or header_pos")

# ====================================================================
# Function    : get_index_in_col_header(sheet, item_name, header_pos = 1)
# Description : get the index of the item in the header
# ====================================================================
def get_index_in_col_header (sheet, item_name, header_pos = 1):
    index    = -1
    max_row  = sheet.max_row
    col_list = get_row_from_sheet(sheet, header_pos, max_row, 1)
    for _ in range(0, len(col_list)):
        if (item_name == col_list[_]):
            index = _ + 1
            return index
    print("The item: " + item_name + " is not found in the header, please check the item name or header_pos")


